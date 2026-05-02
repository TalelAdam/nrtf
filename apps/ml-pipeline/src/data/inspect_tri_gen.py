"""
inspect_tri_gen.py
==================

Step 1 of the cleaning pipeline.

WHY THIS EXISTS
---------------
Before we write any "real" cleaning code, we need to know what the raw data
actually looks like. Industrial Excel reports are notoriously messy — headers
can start on row 5, dates can live in a sidebar, units can be hidden in the
column name, sheets can be named "Feuil1" through "Feuil12", and so on.
Cleaning blind always wastes time.

This script does NOT transform anything. It only OBSERVES and REPORTS.

WHAT IT DOES
------------
1.  Walks the raw tri-gen folder, listing every .xlsx file (we ignore .pdf).
2.  For each workbook, opens every sheet with openpyxl in read-only mode.
3.  Prints/saves a markdown report containing, per sheet:
      - file name + sheet name
      - dimensions (rows × cols)
      - the first 12 rows of values, formatted as a table
      - the script's *guess* at:
          * which row is the header
          * which column is the date / timestamp
          * which columns are numeric "sensor" measurements
          * detected units (kWh, MWh, m³, °C, bar, t, kg, Nm³, ...)
4.  Aggregates a "global sensor list" across files (so we can see
    name variants like "Énergie électrique" vs "Energie elec." vs "EE").
5.  Flags suspicious files (looks like a duplicate, e.g. "... (1).xlsx").

OUTPUT
------
- A pretty terminal log.
- D:/Hackathons/NRTF3/repo/nrtf/data/processed/tri-gen/_inspection_report.md
  (regenerable; safe to delete and re-run).

USAGE
-----
    cd apps/ml-pipeline
    python -m venv .venv
    .venv\\Scripts\\activate          # Windows
    # source .venv/bin/activate        # macOS/Linux
    pip install -r requirements.txt
    python -m src.data.inspect_tri_gen

Then paste the contents of _inspection_report.md back to me (Claude) and
we'll write the real cleaner around what's actually there.
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Any

# These imports are intentionally cheap. If openpyxl is missing the user gets a
# crisp error pointing them at requirements.txt.
try:
    from openpyxl import load_workbook
except ImportError as e:  # pragma: no cover
    print("openpyxl is required. Run: pip install -r requirements.txt", file=sys.stderr)
    raise

# --------------------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------------------

# Default raw-data location. Overridden by --input on the CLI.
# Note the user has the data outside the repo for now; we point at it directly.
DEFAULT_RAW_DIR = Path(
    r"D:\Hackathons\NRTF3\nrtf data -20260501T231459Z-3-001\nrtf data\data tri gen"
)

# Where to drop the report.
REPO_ROOT = Path(__file__).resolve().parents[4]  # apps/ml-pipeline/src/data/inspect_tri_gen.py
DEFAULT_REPORT_PATH = REPO_ROOT / "data" / "processed" / "tri-gen" / "_inspection_report.md"

# How many leading rows of each sheet to render in the report. Industrial reports
# often have 4–6 rows of header above the data; 12 is enough to see the structure.
PREVIEW_ROWS = 12

# Patterns that look like an energy/industrial unit. Used both for sensor-column
# detection AND for unit extraction from header strings like "Gaz (Nm³)".
UNIT_PATTERN = re.compile(
    r"\b("
    r"kwh|mwh|gwh|"           # electrical / thermal energy
    r"kj|mj|gj|"              # energy (joules)
    r"kw|mw|"                 # power
    r"m\^?3|nm\^?3|m³|nm³|"   # gas/water volume (normal-cubic-metre too)
    r"l|m3/h|l/h|l/min|"      # flows
    r"bar|mbar|pa|kpa|"       # pressure
    r"°c|degc|"               # temperature
    r"kg|t|tonne|"            # mass
    r"%|"                     # ratios
    r"kgco2|tco2|kgco2eq"     # CO₂
    r")\b",
    re.IGNORECASE,
)

# Hints that a header string is the date column.
DATE_HEADER_HINTS = {"date", "jour", "day", "heure", "hour", "datetime", "horodatage", "timestamp"}

# Files we skip outright (binary previews, etc.).
SKIP_SUFFIXES = {".pdf", ".png", ".jpg", ".jpeg", ".gif"}

# --------------------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------------------


def cell_to_jsonable(v: Any) -> Any:
    """Coerce an openpyxl cell value into something JSON / markdown-safe."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, float) and v != v:  # NaN
        return ""
    return v


def looks_like_date(v: Any) -> bool:
    """Cheap test: is this cell already a date-typed value, or a date-shaped string?"""
    if isinstance(v, (datetime, date)):
        return True
    if isinstance(v, str):
        s = v.strip()
        # Common French/European date shapes: 01/04/2026, 2026-04-01, 1-avr-26 ...
        return bool(re.match(r"^\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}", s))
    return False


def detect_unit(s: str) -> str | None:
    """Pull a unit token out of a header label like 'Gaz consommé (Nm³)'."""
    if not isinstance(s, str):
        return None
    m = UNIT_PATTERN.search(s)
    return m.group(1) if m else None


def is_numeric(v: Any) -> bool:
    if isinstance(v, (int, float)) and not (isinstance(v, float) and v != v):
        return True
    if isinstance(v, str):
        # Tolerate French decimals "1 234,56"
        cleaned = v.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        try:
            float(cleaned)
            return True
        except ValueError:
            return False
    return False


def guess_header_row(rows: list[list[Any]]) -> int:
    """
    Heuristic: the header row is the *last* row before numeric data starts that
    has many string cells. We scan top to bottom and pick the first row where:
      - the row below is mostly numeric, AND
      - this row has >= 2 non-empty string cells.
    Falls back to row 0.
    """
    for i in range(min(len(rows) - 1, 15)):
        cur, nxt = rows[i], rows[i + 1]
        cur_strs = sum(1 for c in cur if isinstance(c, str) and c.strip())
        nxt_nums = sum(1 for c in nxt if is_numeric(c))
        if cur_strs >= 2 and nxt_nums >= 2:
            return i
    return 0


def guess_date_column(header: list[Any], data_rows: list[list[Any]]) -> int | None:
    """Prefer header-name match; fall back to "the column with the most date-like cells"."""
    # 1) Header hint
    for idx, h in enumerate(header):
        if isinstance(h, str) and h.strip().lower() in DATE_HEADER_HINTS:
            return idx
    # 2) Cell shape vote
    if not data_rows:
        return None
    width = max(len(r) for r in data_rows)
    votes = [0] * width
    for r in data_rows[:50]:
        for i, c in enumerate(r):
            if i < width and looks_like_date(c):
                votes[i] += 1
    best = max(range(width), key=lambda i: votes[i])
    return best if votes[best] > 0 else None


def md_escape(s: Any) -> str:
    """Escape pipe + newline so a value renders cleanly in a markdown table cell."""
    s = str(cell_to_jsonable(s))
    return s.replace("|", "\\|").replace("\n", " ").replace("\r", " ")[:60]


# --------------------------------------------------------------------------------------
# Per-sheet inspection
# --------------------------------------------------------------------------------------


def inspect_sheet(file: Path, sheet_name: str, ws) -> dict:
    rows: list[list[Any]] = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
        if len(rows) >= 200:  # don't pull the whole file; we just want shape
            break

    # Trim trailing all-empty rows from the preview window
    def is_empty(row): return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)
    while rows and is_empty(rows[-1]):
        rows.pop()

    if not rows:
        return {
            "file": file.name, "sheet": sheet_name,
            "rows": 0, "cols": 0,
            "header_row_guess": None, "date_col_guess": None,
            "sensor_columns": [], "preview": [],
        }

    n_cols = max(len(r) for r in rows)
    rows = [r + [None] * (n_cols - len(r)) for r in rows]

    header_idx = guess_header_row(rows)
    header = rows[header_idx]
    body = rows[header_idx + 1:]

    date_col = guess_date_column(header, body)

    sensor_columns = []
    for i, h in enumerate(header):
        if i == date_col:
            continue
        label = str(h).strip() if h else ""
        if not label:
            continue
        # A "sensor column" is one where the data rows are mostly numeric.
        col_vals = [body[r][i] for r in range(min(len(body), 30)) if i < len(body[r])]
        n_num = sum(1 for v in col_vals if is_numeric(v))
        if col_vals and n_num >= max(3, len(col_vals) // 3):
            sensor_columns.append({
                "col_index": i,
                "label": label,
                "unit": detect_unit(label),
            })

    return {
        "file": file.name,
        "sheet": sheet_name,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "header_row_guess": header_idx,
        "date_col_guess": date_col,
        "sensor_columns": sensor_columns,
        "preview": [[cell_to_jsonable(c) for c in r] for r in rows[:PREVIEW_ROWS]],
    }


# --------------------------------------------------------------------------------------
# Markdown report
# --------------------------------------------------------------------------------------


def render_markdown(report: list[dict], duplicates: list[str], aggregate: dict) -> str:
    out: list[str] = []
    out.append("# Tri-gen raw data — inspection report")
    out.append("")
    out.append(f"_Generated: {datetime.now().isoformat(timespec='seconds')}_")
    out.append("")
    out.append(f"- Files inspected: **{len({s['file'] for s in report})}**")
    out.append(f"- Sheets inspected: **{len(report)}**")
    out.append(f"- Distinct sensor labels seen: **{len(aggregate['sensor_label_counts'])}**")
    out.append("")

    if duplicates:
        out.append("## Suspected duplicate files")
        out.append("")
        out.append("These look like browser-downloaded copies (e.g. `... (1).xlsx`). "
                   "Decide whether to keep or drop before cleaning.")
        out.append("")
        for d in duplicates:
            out.append(f"- `{d}`")
        out.append("")

    out.append("## Aggregate sensor labels (across all files)")
    out.append("")
    out.append("| Label | Times seen | Unit guess |")
    out.append("|---|---:|---|")
    for label, info in sorted(aggregate["sensor_label_counts"].items(),
                              key=lambda kv: (-kv[1]["count"], kv[0])):
        out.append(f"| {md_escape(label)} | {info['count']} | {info['unit'] or ''} |")
    out.append("")

    out.append("## Per-sheet detail")
    out.append("")
    for s in report:
        out.append(f"### `{s['file']}` :: `{s['sheet']}`")
        out.append("")
        out.append(f"- shape: {s['rows']} rows × {s['cols']} cols")
        out.append(f"- header row guess: **{s['header_row_guess']}**")
        out.append(f"- date column guess: **{s['date_col_guess']}**")
        if s["sensor_columns"]:
            out.append("- sensor columns:")
            for sc in s["sensor_columns"]:
                u = f" _({sc['unit']})_" if sc["unit"] else ""
                out.append(f"  - col {sc['col_index']}: **{sc['label']}**{u}")
        else:
            out.append("- sensor columns: _none detected — needs manual review_")
        out.append("")
        if s["preview"]:
            width = max(len(r) for r in s["preview"])
            out.append("| " + " | ".join(f"c{i}" for i in range(width)) + " |")
            out.append("|" + "---|" * width)
            for row in s["preview"]:
                row = list(row) + [""] * (width - len(row))
                out.append("| " + " | ".join(md_escape(c) for c in row) + " |")
            out.append("")
    return "\n".join(out)


# --------------------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------------------


def main(raw_dir: Path = DEFAULT_RAW_DIR, report_path: Path = DEFAULT_REPORT_PATH) -> None:
    if not raw_dir.exists():
        sys.exit(f"[fatal] raw dir not found: {raw_dir}")

    files = sorted([p for p in raw_dir.iterdir()
                    if p.is_file() and p.suffix.lower() not in SKIP_SUFFIXES])
    xlsx_files = [p for p in files if p.suffix.lower() == ".xlsx"]
    other_files = [p for p in files if p.suffix.lower() != ".xlsx"]

    print(f"[info] raw dir: {raw_dir}")
    print(f"[info] xlsx files: {len(xlsx_files)}; non-xlsx files: {len(other_files)}")

    # Suspect duplicates: foo (1).xlsx, foo (2).xlsx ...
    dup_pat = re.compile(r"\s\(\d+\)(?=\.[^.]+$)")
    duplicates = [p.name for p in xlsx_files if dup_pat.search(p.name)]

    report: list[dict] = []
    aggregate: dict = {"sensor_label_counts": {}}

    for f in xlsx_files:
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            print(f"[warn] could not open {f.name}: {e}")
            continue
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                summary = inspect_sheet(f, sheet_name, ws)
            except Exception as e:
                print(f"[warn] {f.name} :: {sheet_name} failed: {e}")
                continue
            report.append(summary)
            for sc in summary["sensor_columns"]:
                slot = aggregate["sensor_label_counts"].setdefault(
                    sc["label"], {"count": 0, "unit": sc["unit"]}
                )
                slot["count"] += 1
                # Keep the first non-empty unit guess we saw
                if not slot["unit"] and sc["unit"]:
                    slot["unit"] = sc["unit"]
        wb.close()

    md = render_markdown(report, duplicates, aggregate)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(md, encoding="utf-8")
    json_path = report_path.with_suffix(".json")
    json_path.write_text(json.dumps({"report": report, "duplicates": duplicates,
                                     "aggregate": aggregate}, indent=2,
                                    default=str, ensure_ascii=False), encoding="utf-8")

    print(f"[ok] wrote markdown report  -> {report_path}")
    print(f"[ok] wrote machine-readable -> {json_path}")
    print()
    print("Next: open the markdown report, sanity-check the header/date guesses,")
    print("then paste the 'Aggregate sensor labels' section back to Claude so we")
    print("can lock the cleaning schema.")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Inspect tri-gen Excel reports.")
    parser.add_argument("--input", type=Path, default=DEFAULT_RAW_DIR,
                        help="Folder containing the raw .xlsx monthly reports.")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_PATH,
                        help="Path to write the markdown report to.")
    args = parser.parse_args()
    main(args.input, args.report)
