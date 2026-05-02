"""
clean_tri_gen.py
================

Step 2: take the raw monthly Excel reports for the tri-gen plant, normalize them
into a single canonical "long" Parquet table, and emit one Excel per sensor for
the chem/bio teammates.

WHY LONG (TIDY) FORMAT
----------------------
Wide format ("one column per sensor") is convenient for humans but painful for
ML: every model has to know the schema, joins are hard, and adding a new sensor
requires altering the table. Long format — (timestamp, sensor_id, value, unit) —
makes the rest of the pipeline trivial:

    polars.scan_parquet(...)                  # one read
        .filter(sensor_id == "elec_total")    # one model per sensor, same code
        .group_by_dynamic("ts", every="1h")
        ...

WHAT THIS SCRIPT EXPECTS
------------------------
Run `inspect_tri_gen.py` first. It writes `_inspection_report.json` next to the
markdown report. This cleaner reads that JSON to learn each sheet's
(header_row, date_column, sensor_columns) — so we don't re-guess.

If a sheet's guesses are wrong, edit `_inspection_report.json` (or add an entry
to the OVERRIDES dict below) and re-run.

OUTPUT
------
data/processed/tri-gen/
├── long.parquet                 # canonical: ts, sensor_id, value, unit, source_file, source_sheet, source_row
├── by-sensor/
│   ├── elec_total.xlsx          # one workbook per sensor, sorted by ts
│   ├── gaz_consomme.xlsx
│   └── ...
├── _build_manifest.json         # SHAs + row counts + build timestamp
└── _cleaning_log.txt            # what got dropped/coerced and why
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    print("Run: pip install -r requirements.txt", file=sys.stderr)
    raise

REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_RAW_DIR = Path(
    r"D:\Hackathons\NRTF3\nrtf data -20260501T231459Z-3-001\nrtf data\data tri gen"
)
DEFAULT_OUT_DIR = REPO_ROOT / "data" / "processed" / "tri-gen"
DEFAULT_REPORT_JSON = DEFAULT_OUT_DIR / "_inspection_report.json"

# Sensor-label normalisation. After we run the inspector we'll see the real
# variants; for now this dict is empty and we just slugify whatever we see.
# Example you might add later:
#   "Énergie électrique totale": "elec_total",
#   "Energie elec.":              "elec_total",
LABEL_TO_SENSOR_ID: dict[str, str] = {}

# Per-sheet manual overrides. Use only when the inspector's guess is wrong.
# Key is "filename::sheetname"; value can override any field of the inspection.
OVERRIDES: dict[str, dict] = {
    # "FIN OCTOBER _4112025.xlsx::Feuil1": {"header_row_guess": 3, "date_col_guess": 1},
}

UNIT_PATTERN = re.compile(
    r"\b(kwh|mwh|gwh|kw|mw|kvarh|m\^?3|nm\^?3|m³|nm³|m3/h|l/h|bar|°c|degc|kg|t|rpm|v|a|%)\b",
    re.IGNORECASE,
)

# --------------------------------------------------------------------------------------
# Utilities
# --------------------------------------------------------------------------------------

def slugify(s: str) -> str:
    """ASCII-safe lowercase slug. 'Énergie électrique (kWh)' -> 'energie_electrique_kwh'."""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-zA-Z0-9]+", "_", s).strip("_").lower()
    return s or "unknown"


def coerce_number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None  # filter NaN
    if isinstance(v, str):
        s = v.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        if s in {"-", "—", "n/a", "na", "nd"}:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def coerce_datetime(v: Any) -> datetime | None:
    if isinstance(v, datetime):
        return v
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M",
                    "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        try:
            # last-ditch: pandas will accept many shapes
            return pd.to_datetime(s, dayfirst=True, errors="raise").to_pydatetime()
        except Exception:
            return None
    return None


def coerce_time(v: Any) -> tuple[int, int, int] | None:
    """Coerce Excel time cells / strings into (hour, minute, second)."""
    if v is None or v == "":
        return None
    if hasattr(v, "hour") and hasattr(v, "minute") and hasattr(v, "second"):
        return int(v.hour), int(v.minute), int(v.second)
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                t = datetime.strptime(s, fmt)
                return t.hour, t.minute, t.second
            except ValueError:
                continue
    return None


def detect_unit(label: str) -> str:
    m = UNIT_PATTERN.search(label or "")
    return m.group(1) if m else ""


def combine_date_time(date_value: Any, time_value: Any) -> datetime | None:
    d = coerce_datetime(date_value)
    if d is None:
        return None
    t = coerce_time(time_value)
    if t is None:
        return datetime(d.year, d.month, d.day)
    return datetime(d.year, d.month, d.day, t[0], t[1], t[2])


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def find_transposed_date_time_rows(rows: list[list[Any]]) -> tuple[int, int] | None:
    """
    The tri-gen exports are often transposed:

      row "Date"  -> timestamps across columns
      row "Heure" -> times across columns
      later rows  -> one measurement per row

    Return zero-based (date_row, time_row) when that shape is detected.
    """
    for i, row in enumerate(rows[:40]):
        text_cells = [str(c).strip().lower() for c in row[:8] if c not in (None, "")]
        has_date_label = any(c == "date" or c.startswith("date ") for c in text_cells)
        date_votes = sum(1 for c in row if coerce_datetime(c) is not None)
        if not has_date_label and date_votes < 3:
            continue

        best_time_row = i + 1
        best_votes = -1
        for j in range(i + 1, min(i + 4, len(rows))):
            votes = sum(1 for c in rows[j] if coerce_time(c) is not None)
            text = " ".join(str(c).strip().lower() for c in rows[j][:8] if c not in (None, ""))
            if "heure" in text:
                votes += 10
            if votes > best_votes:
                best_time_row = j
                best_votes = votes

        if best_votes > 0:
            return i, best_time_row
    return None


def clean_transposed_sheet(file: Path, sheet_meta: dict, rows: list[list[Any]], log: list[str]) -> list[dict]:
    """Clean sheets where timestamps are columns and measurements are rows."""
    key = f"{file.name}::{sheet_meta['sheet']}"
    found = find_transposed_date_time_rows(rows)
    if found is None:
        log.append(f"[skip] {key}: transposed layout not detected")
        return []

    date_row_idx, time_row_idx = found
    date_row = rows[date_row_idx]
    time_row = rows[time_row_idx]
    width = max(len(r) for r in rows)

    timestamps: dict[int, datetime] = {}
    last_date: Any = None
    for col_idx in range(width):
        if col_idx < len(date_row) and coerce_datetime(date_row[col_idx]) is not None:
            last_date = date_row[col_idx]
        time_value = time_row[col_idx] if col_idx < len(time_row) else None
        ts = combine_date_time(last_date, time_value)
        if ts is not None and coerce_time(time_value) is not None:
            timestamps[col_idx] = ts

    if not timestamps:
        log.append(f"[skip] {key}: no usable timestamp columns found")
        return []

    out: list[dict] = []
    current_group = ""
    for r_idx in range(time_row_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or all(c is None or c == "" for c in row):
            continue

        group_cell = str(row[0]).strip() if len(row) > 0 and row[0] not in (None, "") else ""
        metric_cell = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else ""

        if group_cell:
            current_group = group_cell
        if not metric_cell:
            continue

        label = f"{current_group} - {metric_cell}" if current_group else metric_cell
        sensor_id = LABEL_TO_SENSOR_ID.get(label, slugify(label))
        unit = detect_unit(metric_cell) or detect_unit(current_group)

        values_written = 0
        for col_idx, ts in timestamps.items():
            if col_idx >= len(row):
                continue
            val = coerce_number(row[col_idx])
            if val is None:
                continue
            out.append({
                "ts": ts,
                "sensor_id": sensor_id,
                "sensor_label": label,
                "value": val,
                "unit": unit,
                "source_file": file.name,
                "source_sheet": sheet_meta["sheet"],
                "source_row": r_idx,
            })
            values_written += 1

        if values_written == 0:
            log.append(f"[drop] {key} row={r_idx}: metric {metric_cell!r} had no numeric timestamped values")

    log.append(
        f"[ok-transposed] {key}: date_row={date_row_idx} time_row={time_row_idx} "
        f"timestamp_cols={len(timestamps)} rows={len(out)}"
    )
    return out


# --------------------------------------------------------------------------------------
# Cleaning
# --------------------------------------------------------------------------------------

def clean_one_sheet(file: Path, sheet_meta: dict, log: list[str]) -> list[dict]:
    """Return a list of long-format rows for one sheet, using inspector hints."""
    key = f"{file.name}::{sheet_meta['sheet']}"
    sheet_meta = {**sheet_meta, **OVERRIDES.get(key, {})}

    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb[sheet_meta["sheet"]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    if sheet_meta.get("date_col_guess") is None or not sheet_meta.get("sensor_columns"):
        return clean_transposed_sheet(file, sheet_meta, rows, log)

    header_idx = sheet_meta["header_row_guess"]
    date_col = sheet_meta["date_col_guess"]
    sensor_cols = sheet_meta["sensor_columns"]

    out: list[dict] = []
    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or all(c is None or c == "" for c in row):
            continue
        ts = coerce_datetime(row[date_col]) if date_col < len(row) else None
        if ts is None:
            log.append(f"[drop] {key} row={r_idx}: unparseable date {row[date_col]!r}")
            continue
        for sc in sensor_cols:
            i = sc["col_index"]
            if i >= len(row):
                continue
            val = coerce_number(row[i])
            if val is None:
                continue
            label = sc["label"]
            sensor_id = LABEL_TO_SENSOR_ID.get(label, slugify(label))
            out.append({
                "ts": ts,
                "sensor_id": sensor_id,
                "sensor_label": label,
                "value": val,
                "unit": sc.get("unit") or "",
                "source_file": file.name,
                "source_sheet": sheet_meta["sheet"],
                "source_row": r_idx,
            })
    return out


def main(raw_dir: Path, out_dir: Path, report_json: Path, drop_duplicate_files: bool) -> None:
    if not report_json.exists():
        sys.exit(f"[fatal] inspection report not found: {report_json}\n"
                 f"        run `python -m src.data.inspect_tri_gen` first")

    inspection = json.loads(report_json.read_text(encoding="utf-8"))
    by_file: dict[str, list[dict]] = defaultdict(list)
    for s in inspection["report"]:
        by_file[s["file"]].append(s)

    skipped_dups: list[str] = []
    if drop_duplicate_files:
        skipped_dups = list(inspection.get("duplicates", []))

    out_dir.mkdir(parents=True, exist_ok=True)
    log: list[str] = []
    all_rows: list[dict] = []
    for fname, sheets in by_file.items():
        if fname in skipped_dups:
            log.append(f"[skip-dup] {fname}")
            continue
        f = raw_dir / fname
        if not f.exists():
            log.append(f"[miss] {fname} listed in inspection but not found on disk")
            continue
        for sheet_meta in sheets:
            all_rows.extend(clean_one_sheet(f, sheet_meta, log))

    if not all_rows:
        sys.exit("[fatal] no rows produced — review _cleaning_log.txt")

    df = pd.DataFrame(all_rows)
    df = df.sort_values(["sensor_id", "ts"]).reset_index(drop=True)

    # Deduplicate exact (ts, sensor_id) collisions (e.g. when the same month
    # appears in two reports). Keep the first; log the rest.
    before = len(df)
    df = df.drop_duplicates(subset=["ts", "sensor_id"], keep="first")
    log.append(f"[dedup] dropped {before - len(df)} duplicate (ts, sensor_id) rows")

    # 1) Canonical long parquet
    long_path = out_dir / "long.parquet"
    df.to_parquet(long_path, index=False)
    log.append(f"[ok] wrote canonical long table: {long_path}  (rows={len(df)})")

    # 2) Per-sensor Excel exports
    by_sensor_dir = out_dir / "by-sensor"
    by_sensor_dir.mkdir(parents=True, exist_ok=True)
    for sensor_id, group in df.groupby("sensor_id"):
        path = by_sensor_dir / f"{sensor_id}.xlsx"
        with pd.ExcelWriter(path, engine="xlsxwriter", datetime_format="yyyy-mm-dd hh:mm:ss") as xw:
            group[["ts", "value", "unit", "sensor_label",
                   "source_file", "source_sheet", "source_row"]].to_excel(
                xw, sheet_name="data", index=False)
        log.append(f"[ok] wrote sensor file: {path}  (rows={len(group)})")

    # 3) Manifest
    manifest = {
        "name": "tri_gen_v1",
        "version": "1.0.0",
        "build_ts": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "row_count": int(len(df)),
        "sensor_count": int(df["sensor_id"].nunique()),
        "sensors": sorted(df["sensor_id"].unique().tolist()),
        "ts_min": df["ts"].min().isoformat(),
        "ts_max": df["ts"].max().isoformat(),
        "source_files": sorted(df["source_file"].unique().tolist()),
        "source_file_shas": {
            p.name: file_sha256(p)
            for p in sorted(raw_dir.iterdir())
            if p.suffix.lower() == ".xlsx" and p.name not in skipped_dups
        },
    }
    (out_dir / "_build_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    (out_dir / "_cleaning_log.txt").write_text("\n".join(log), encoding="utf-8")
    print("\n".join(log[-30:]))
    print(f"\n[done] {len(df)} rows / {df['sensor_id'].nunique()} sensors -> {out_dir}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Clean tri-gen Excel reports to canonical Parquet.")
    parser.add_argument("--input", type=Path, default=DEFAULT_RAW_DIR)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_JSON)
    parser.add_argument("--keep-duplicates", action="store_true",
                        help="Keep '... (N).xlsx' duplicate files instead of dropping.")
    args = parser.parse_args()
    main(args.input, args.out, args.report, drop_duplicate_files=not args.keep_duplicates)
